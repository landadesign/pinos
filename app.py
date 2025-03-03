import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
import re
import openpyxl

# 定数
RATE_PER_KM = 15
DAILY_ALLOWANCE = 200

# 計算日付の設定
calculationDate = datetime.now().strftime("%Y%m%d")

def create_expense_table_image(df, name):
    # 画像サイズとフォント設定
    width = 1200
    row_height = 40
    header_height = 60
    padding = 30
    title_height = 50
    
    # 全行数を計算（タイトル + ヘッダー + データ行 + 合計行 + 注釈）
    total_rows = len(df) + 3
    height = title_height + header_height + (total_rows * row_height) + padding * 2
    
    # 画像作成
    img = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(img)
    
    # フォント設定
    font = ImageFont.load_default()
    
    # タイトル描画
    title = f"{name}様 2024年12月25日～2025年1月 社内通貨（交通費）清算額"
    draw.text((padding, padding), title, fill='black', font=font)
    
    # ヘッダー
    headers = ['日付', '経路', '合計距離(km)', '交通費（距離×15P）(円)', '運転手当(円)', '合計(円)']
    x_positions = [padding, padding + 80, padding + 600, padding + 750, padding + 900, padding + 1050]
    
    y = padding + title_height
    for header, x in zip(headers, x_positions):
        draw.text((x, y), header, fill='black', font=font)
    
    # 罫線
    line_y = y + header_height - 5
    draw.line([(padding, line_y), (width - padding, line_y)], fill='black', width=1)
    
    # データ行
    y = padding + title_height + header_height
    for _, row in df.iterrows():
        for route in row['routes']:
            # 日付
            draw.text((x_positions[0], y), str(row['date']), fill='black', font=font)
            
            # 経路
            draw.text((x_positions[1], y), route['route'], fill='black', font=font)
            
            # 最初のルートの行にのみ数値を表示
            if route == row['routes'][0]:
                # 距離
                draw.text((x_positions[2], y), f"{row['total_distance']:.1f}", fill='black', font=font)
                
                # 交通費
                draw.text((x_positions[3], y), f"{int(row['transportation_fee']):,}", fill='black', font=font)
                
                # 運転手当
                draw.text((x_positions[4], y), f"{int(row['allowance']):,}", fill='black', font=font)
                
                # 合計
                draw.text((x_positions[5], y), f"{int(row['total']):,}", fill='black', font=font)
            
            y += row_height
    
    # 合計行の罫線
    line_y = y - 5
    draw.line([(padding, line_y), (width - padding, line_y)], fill='black', width=1)
    
    # 合計行
    y += 10
    draw.text((x_positions[0], y), "合計", fill='black', font=font)
    draw.text((x_positions[5], y), f"{int(df['total'].sum()):,}", fill='black', font=font)
    
    # 注釈
    y += row_height * 2
    draw.text((padding, y), "※2025年1月分給与にて清算しました。", fill='black', font=font)
    
    # 計算日時
    draw.text((padding, y + row_height), f"計算日時: {datetime.now().strftime('%Y/%m/%d')}", fill='black', font=font)
    
    # 画像をバイト列に変換
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()
    
    return img_byte_arr

def parse_expense_data(text):
    """テキストデータを解析してDataFrameを作成"""
    lines = text.replace('\r\n', '\n').split('\n')
    data = []
    entry_id = 1
    current_entry = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # 新しいエントリーの開始を検出
        if '【ピノ】' in line:
            # 前のエントリーを処理
            if current_entry:
                entry_data = process_entry(''.join(current_entry))
                if entry_data:
                    entry_data['id'] = entry_id
                    data.append(entry_data)
                    entry_id += 1
            current_entry = [line]
        else:
            # 距離情報を含む行を追加
            if any(pattern in line for pattern in ['km', '㎞', 'ｋｍ', 'kｍ']):
                current_entry.append(line)
    
    # 最後のエントリーを処理
    if current_entry:
        entry_data = process_entry(''.join(current_entry))
        if entry_data:
            entry_data['id'] = entry_id
            data.append(entry_data)
    
    return pd.DataFrame(data)

def process_entry(text):
    """個別のエントリーを解析"""
    # 基本情報の抽出（名前と日付）
    name_date_match = re.search(r'【ピノ】\s*([^　\s]+(?:[ 　]+[^　\s]+)*)\s+(\d+/\d+)\s*\([月火水木金土日]\)', text)
    if not name_date_match:
        return None
        
    name = name_date_match.group(1).strip()
    date = name_date_match.group(2)
    
    # 距離の抽出（複数のパターンに対応）
    distance_patterns = [
        r'(\d+\.?\d*)(?:km|㎞|ｋｍ|kｍ)',  # 基本パターン
        r'距離[：:]\s*(\d+\.?\d*)',        # 「距離:」パターン
        r'合計[：:]\s*(\d+\.?\d*)',        # 「合計:」パターン
        r'往復[：:]\s*(\d+\.?\d*)',        # 「往復:」パターン
    ]
    
    distance = None
    for pattern in distance_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            distance = float(match.group(1))
            break
    
    if not distance:
        return None
    
    # 経路の抽出（日付の後ろから距離の前まで）
    route_text = text[text.find(')')+1:text.find(str(distance))].strip()
    route = re.sub(r'\s+', ' ', route_text)  # 複数の空白を1つに
    
    return {
        'name': name,  # 【ピノ】の後の名前を使用
        'date': date,
        'route': route,
        'distance': distance
    }

def create_expense_report(person_data):
    """個人の精算書データを作成（経路ごとに表示）"""
    # データを日付順にソート
    person_data = person_data.sort_values('date')
    
    # 各経路のデータを作成
    route_data = []
    current_date = None
    
    for _, row in person_data.iterrows():
        # 日付が変わったら運転手当を付与
        if current_date != row['date']:
            current_date = row['date']
            add_allowance = 200  # 新しい日付の最初の経路に運転手当を付与
        else:
            add_allowance = 0  # 同じ日付の2つ目以降の経路には運転手当を付与しない
        
        route_data.append({
            '日付': row['date'],
            '経路': row['route'],
            '合計距離(km)': round(row['distance'], 1),
            '交通費（距離×15P）(円)': int(row['distance'] * 15),
            '運転手当(円)': add_allowance,
            '合計(円)': int(row['distance'] * 15 + add_allowance)
        })
    
    # DataFrameを作成
    expense_data = pd.DataFrame(route_data)
    
    # 合計行の追加
    total_row = pd.DataFrame({
        '日付': ['合計'],
        '経路': [''],
        '合計距離(km)': [expense_data['合計距離(km)'].sum()],
        '交通費（距離×15P）(円)': [expense_data['交通費（距離×15P）(円)'].sum()],
        '運転手当(円)': [expense_data['運転手当(円)'].sum()],
        '合計(円)': [expense_data['合計(円)'].sum()]
    })
    
    # データフレームを結合
    expense_data = pd.concat([expense_data, total_row], ignore_index=True)
    
    return expense_data

def create_pdf(expense_data, name):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # データをリスト形式に変換
    data = [expense_data.columns.tolist()]  # ヘッダー
    data.extend(expense_data.values.tolist())  # データ行
    
    # テーブルスタイルの設定
    table = Table(data)
    style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ])
    table.setStyle(style)
    
    # タイトルと注釈を追加
    elements = []
    elements.append(table)
    
    # PDFを生成
    doc.build(elements)
    
    return buffer

def create_png(expense_data, name):
    # HTMLテーブルを作成
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; }}
            table {{ 
                border-collapse: collapse; 
                width: 100%;
                margin-bottom: 20px;
            }}
            th, td {{ 
                border: 1px solid #ddd; 
                padding: 8px; 
                text-align: left; 
            }}
            th {{ 
                background-color: #2196F3; 
                color: white; 
            }}
            td {{ text-align: right; }}
            td:nth-child(1), td:nth-child(2) {{ text-align: left; }}
            .title {{ 
                text-align: center; 
                margin-bottom: 20px; 
            }}
            .note {{ 
                margin-top: 15px; 
                color: #666; 
                font-size: 0.9em; 
            }}
        </style>
    </head>
    <body>
        <h2 class="title">{name}様 2024年12月25日～2025年1月 社内通貨（交通費）清算額</h2>
        {expense_data.to_html(index=False)}
        <div class="note">※2025年1月分給与にて清算しました。</div>
    </body>
    </html>
    """
    
    # Chromeのオプション設定
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # ヘッドレスモード
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    
    # WebDriverの初期化
    driver = webdriver.Chrome(options=chrome_options)
    
    try:
        # HTMLの読み込みとスクリーンショット
        driver.get("data:text/html;charset=utf-8," + html_content)
        driver.set_window_size(1200, len(expense_data) * 30 + 200)
        png_data = driver.get_screenshot_as_png()
        return png_data
    finally:
        driver.quit()

def capture_streamlit_table():
    # 少し待って画面が描画されるのを待つ
    time.sleep(1)
    # 画面全体のスクリーンショットを取得
    screenshot = pyautogui.screenshot()
    return screenshot

def export_to_excel(df, unique_names):
    """精算書をExcelファイルとして出力"""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    
    # 最初のシートを使用
    first_name = unique_names[0]
    first_sheet = workbook.active
    first_sheet.title = f"{first_name}様"
    
    # 全ての担当者のシートを作成
    for name in unique_names:
        if name == first_name:
            worksheet = first_sheet
        else:
            worksheet = workbook.create_sheet(f"{name}様")
        
        # 担当者のデータを抽出して精算書を作成
        person_data = df[df['name'] == name].copy()
        expense_data = create_expense_report(person_data)
        
        # A4サイズに合わせた設定
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        
        # 列幅の設定
        worksheet.column_dimensions['A'].width = 15  # 日付
        worksheet.column_dimensions['B'].width = 50  # 経路
        worksheet.column_dimensions['C'].width = 15  # 合計距離
        worksheet.column_dimensions['D'].width = 25  # 交通費
        worksheet.column_dimensions['E'].width = 15  # 運転手当
        worksheet.column_dimensions['F'].width = 15  # 合計
        
        # 行の高さを設定
        worksheet.row_dimensions[1].height = 45  # タイトル行
        worksheet.row_dimensions[2].height = 60  # ヘッダー行
        
        # タイトルを追加
        title = f"{name}様 2025年1月 社内通貨（交通費）清算額"
        worksheet['A1'] = title
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        title_cell.font = openpyxl.styles.Font(size=14, bold=True)
        
        # ヘッダーの書き込み
        headers = ['日付', '経路', '合計距離(km)', '交通費（距離×15P）(円)', '運転手当(円)', '合計(円)']
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(size=11, bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # データの書き込み
        for row_idx, row in enumerate(expense_data.values, 3):
            worksheet.row_dimensions[row_idx].height = 30  # データ行の高さを30に設定
            
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = openpyxl.styles.Font(size=11)
                
                # 数値の書式設定
                if col_idx in [3, 4, 5, 6]:  # 数値列
                    cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
                    if col_idx == 3:  # 距離
                        cell.number_format = '#,##0.0'
                    else:  # 金額
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 罫線の設定
        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(expense_data.values)+2, min_col=1, max_col=6):
            for cell in row:
                cell.border = border
        
        # 注釈を追加
        note_row = len(expense_data) + 4
        worksheet.row_dimensions[note_row].height = 30
        note_cell = worksheet[f'A{note_row}']
        note_cell.value = "※2025年1月分給与にて清算しました。"
        worksheet.merge_cells(f'A{note_row}:F{note_row}')
        note_cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        note_cell.font = openpyxl.styles.Font(size=9)
    
    # ファイルを保存
    workbook.save(output)
    
    return output.getvalue()

def main():
    # ページ設定
    st.set_page_config(
        page_title="PINO精算アプリケーション",
        layout="wide"
    )
    
    st.title("PINO精算アプリケーション")
    
    # テキストエリアの表示
    input_text = st.text_area("精算データを貼り付けてください", height=200)
    
    # 解析ボタンとクリアボタン
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("データを解析"):
            if input_text:
                df = parse_expense_data(input_text)
                st.session_state['df'] = df
                st.success("データを解析しました！")
    with col2:
        if st.button("クリア"):
            st.session_state['df'] = None
            st.session_state['show_expense_report'] = False
            st.rerun()
    
    # データ一覧と精算書の表示
    if 'df' in st.session_state and st.session_state['df'] is not None:
        df = st.session_state['df']
        if not df.empty:
            # データ一覧の表示
            st.markdown("""
            <h2 style='text-align: center; padding: 20px 0;'>
                交通費データ一覧
            </h2>
            """, unsafe_allow_html=True)
            
            st.dataframe(
                df,
                column_config={
                    'id': st.column_config.NumberColumn('No.', width=70),
                    'date': st.column_config.TextColumn('日付', width=100),
                    'name': st.column_config.TextColumn('担当者', width=120),
                    'route': st.column_config.TextColumn('経路', width=500),
                    'distance': st.column_config.NumberColumn('距離(km)', format="%.1f", width=100)
                },
                hide_index=True
            )
            
            # 精算書を表示するボタン
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("精算書を表示"):
                    st.session_state['show_expense_report'] = True
                    st.rerun()
            
            # 精算書の表示
            if st.session_state.get('show_expense_report', False):
                # 担当者ごとのタブを作成
                unique_names = sorted(df['name'].unique())
                tabs = st.tabs([f"{name}様" for name in unique_names])
                
                # 担当者ごとの精算書表示
                for i, name in enumerate(unique_names):
                    with tabs[i]:
                        title = f"{name}様 2025年1月 社内通貨（交通費）清算額"
                        st.markdown(f"### {title}")
                        
                        # 担当者のデータを抽出して精算書を作成
                        person_data = df[df['name'] == name].copy()
                        expense_data = create_expense_report(person_data)
                        
                        # 精算書の表示
                        st.dataframe(
                            expense_data,
                            column_config={
                                '日付': st.column_config.TextColumn('日付', width=120),
                                '経路': st.column_config.TextColumn('経路', width=600),
                                '合計距離(km)': st.column_config.NumberColumn(
                                    '合計距離(km)',
                                    format="%.1f",
                                    width=150
                                ),
                                '交通費（距離×15P）(円)': st.column_config.NumberColumn(
                                    '交通費（距離×15P）(円)',
                                    format="%.0f",
                                    width=200
                                ),
                                '運転手当(円)': st.column_config.NumberColumn(
                                    '運転手当(円)',
                                    format="%.0f",
                                    width=150
                                ),
                                '合計(円)': st.column_config.NumberColumn(
                                    '合計(円)',
                                    format="%.0f",
                                    width=150
                                )
                            },
                            hide_index=True,
                            height=400
                        )
                        
                        # 注釈表示
                        st.markdown("""
                            <div style='margin-top: 15px; color: #666;'>
                                ※2025年1月分給与にて清算しました。
                            </div>
                        """, unsafe_allow_html=True)
                
                # ダウンロードボタン
                st.markdown("---")
                excel_data = export_to_excel(df, unique_names)  # unique_namesはここで定義済み
                st.download_button(
                    label="精算書をExcelでダウンロード",
                    data=excel_data,
                    file_name=f'精算書_2025年1月.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

if __name__ == "__main__":
    main()
